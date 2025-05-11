# Apache Software License 2.0
#
# Copyright (c) ZenML GmbH 2025. All rights reserved.
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
# http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
#

from pathlib import Path
from typing import Annotated, Dict

import pandas as pd
from openpyxl import Workbook, load_workbook
from zenml import get_step_context, log_metadata, step

from utils import score_risk

# --------------------------------------------------------------------------- #
RiskScores = Annotated[Dict[str, float], "risk_scores"]
# --------------------------------------------------------------------------- #


@step
def risk_assessment(evaluation_results: Dict) -> RiskScores:
    """Compute risk scores & update register. Article 9 compliant.

    Converts evaluation metrics + bias flag → quantitative risk score
    Updates risk_register.xlsx (one row / model version)
    Logs summary via `log_metadata`

    Args:
        evaluation_results: Dictionary containing evaluation results.

    Returns:
        Dictionary containing risk scores.
    """
    scores = score_risk(evaluation_results)

    # ---- update Excel risk register ---------------------------------------
    reg_path = Path("compliance/risk_register.xlsx")
    reg_path.parent.mkdir(parents=True, exist_ok=True)

    # Ensure workbook + sheet exist
    if not reg_path.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = "Risks"
        ws.append(["Run_ID", "Risk_overall", "Risk_auc", "Risk_bias", "Status"])
        wb.save(reg_path)

    wb = load_workbook(reg_path)
    ws = wb["Risks"]

    # Get run_id from step context
    run_id = get_step_context().pipeline_run.id

    # Check if this run already logged
    run_ids = [cell.value for cell in ws["A"][1:]]  # skip header
    if run_id in run_ids:
        row_idx = run_ids.index(run_id) + 2
    else:
        row_idx = ws.max_row + 1

    ws.cell(row=row_idx, column=1).value = run_id
    ws.cell(row=row_idx, column=2).value = scores["overall"]
    ws.cell(row=row_idx, column=3).value = scores["risk_auc"]
    ws.cell(row=row_idx, column=4).value = scores["risk_bias"]
    ws.cell(row=row_idx, column=5).value = (
        "Mitigation needed" if scores["overall"] > 0.4 else "Acceptable"
    )

    wb.save(reg_path)

    # ---- export Markdown snapshot -----------------------------------------
    df = pd.read_excel(reg_path, sheet_name="Risks")
    md_path = Path("compliance/manual_fills/risk_register.md")
    md_path.parent.mkdir(parents=True, exist_ok=True)
    md_path.write_text(df.to_markdown(index=False))

    # ---- log metadata ------------------------------------------------------
    log_metadata({"risk_scores": scores})

    return scores
