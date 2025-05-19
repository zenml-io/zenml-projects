# Apache Software License 2.0
#
# Copyright (c) ZenML GmbH 2025. All rights reserved.
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
# http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
#

import os
from datetime import datetime
from pathlib import Path
from typing import Annotated, Dict, List

from openpyxl import Workbook, load_workbook
from zenml import get_step_context, log_metadata, step
from zenml.logger import get_logger

from src.constants import (
    APPROVAL_THRESHOLDS,
    HAZARD_DEFINITIONS,
    RISK_REGISTER_PATH,
    RISK_SCORES_NAME,
)
from src.utils.modal_utils import save_artifact_to_modal

logger = get_logger(__name__)

# --------------------------------------------------------------------------- #
RiskScores = Annotated[Dict[str, float], RISK_SCORES_NAME]
# --------------------------------------------------------------------------- #


def identify_hazards(evaluation_results: Dict, scores: Dict) -> List[Dict]:
    """Identify applicable hazards based on evaluation results."""
    identified_hazards = []

    # Evaluate each hazard trigger function
    for hazard_id, hazard_info in HAZARD_DEFINITIONS.items():
        try:
            trigger_func = hazard_info["trigger"]
            if trigger_func(evaluation_results, scores):
                identified_hazards.append(
                    {
                        "id": hazard_id,
                        "description": hazard_info["description"],
                        "severity": hazard_info["severity"],
                        "mitigation": hazard_info["mitigation"],
                    }
                )
        except (KeyError, TypeError) as e:
            logger.warning(f"Could not evaluate hazard {hazard_id}: {e}")

    return identified_hazards


def score_risk(evaluation: Dict) -> Dict[str, float]:
    """Return dict with per-factor risk ∈ [0,1] + overall."""
    # Get metrics with default fallbacks
    metrics = evaluation.get("metrics", {})
    if not isinstance(metrics, dict):
        metrics = {}

    # Get AUC with default
    auc = metrics.get("auc", 0.5)

    # Get fairness info with defaults
    fairness = evaluation.get("fairness", {})
    if not isinstance(fairness, dict):
        fairness = {}

    bias_flag = fairness.get("bias_flag", False)

    # Calculate max disparity across groups
    disparities = []
    fairness_metrics = fairness.get("fairness_metrics", {})
    if isinstance(fairness_metrics, dict):
        for group_metrics in fairness_metrics.values():
            if isinstance(group_metrics, dict) and "selection_rate_disparity" in group_metrics:
                disparities.append(abs(group_metrics["selection_rate_disparity"]))

    disparity = max(disparities) if disparities else 0.0

    # Calculate risk scores
    risk_auc = 1 - auc  # low AUC → higher risk
    risk_bias = 0.8 if bias_flag else disparity  # flat score if flagged
    overall = round(min(1.0, 0.5 * risk_auc + 0.5 * risk_bias), 3)

    return {
        "risk_auc": round(risk_auc, 3),
        "risk_bias": round(risk_bias, 3),
        "overall": overall,
    }


def get_status(risk_score: float) -> str:
    """Determine row status for the risk register.

    Args:
        risk_score: The overall risk score ∈ [0,1].

    Returns:
        "PENDING" if the score exceeds the configured threshold, otherwise "COMPLETED".
    """
    threshold = APPROVAL_THRESHOLDS["risk_score"]
    return "PENDING" if risk_score > threshold else "COMPLETED"


@step
def risk_assessment(evaluation_results: Dict) -> RiskScores:
    """Compute risk scores & update register. Article 9 compliant.

    Converts evaluation metrics + bias flag → quantitative risk score
    Updates risk_register.xlsx (one row / model version)
    Logs summary via `log_metadata`

    Args:
        evaluation_results: Dictionary containing evaluation results.

    Returns:
        Dictionary containing risk scores.
    """
    scores = score_risk(evaluation_results)
    hazards = identify_hazards(evaluation_results, scores)

    headers = [
        "Run_ID",
        "Timestamp",
        "Risk_overall",
        "Risk_auc",
        "Risk_bias",
        "Risk_category",
        "Status",
        "Risk_description",
        "Mitigation",
    ]

    # Create a fresh workbook
    wb_path = Path(RISK_REGISTER_PATH)
    if wb_path.exists():
        wb = load_workbook(wb_path)
        ws = wb["Risks"]
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Risks"
        ws.append(headers)

    # Get run_id
    run_id = get_step_context().pipeline_run.id
    timestamp = datetime.now().isoformat()

    # Add record for this run
    if hazards:
        # Add to main Risks sheet
        for hz in hazards:
            ws.append(
                [
                    str(run_id),
                    timestamp,
                    scores["overall"],
                    scores["risk_auc"],
                    scores["risk_bias"],
                    hz["severity"].upper(),
                    get_status(scores["overall"]),
                    hz["description"],
                    hz["mitigation"],
                ]
            )

        # Add to HazardDetails sheet
        if "HazardDetails" not in wb.sheetnames:
            hazard_sheet = wb.create_sheet("HazardDetails")
            hazard_sheet.append(
                [
                    "Run_ID",
                    "Timestamp",
                    "Risk_Score",
                    "Hazard_ID",
                    "Description",
                    "Severity",
                    "Mitigation",
                    "Details",
                ]
            )
        else:
            hazard_sheet = wb["HazardDetails"]

        # Add records for each hazard
        for hz in hazards:
            # Get additional information for bias hazards
            details = ""
            if hz["id"] == "bias_protected_groups" and "fairness" in evaluation_results:
                fairness = evaluation_results["fairness"]
                if isinstance(fairness, dict) and "fairness_metrics" in fairness:
                    for attr, metrics in fairness.get("fairness_metrics", {}).items():
                        if isinstance(metrics, dict) and "selection_rate_disparity" in metrics:
                            disparity = metrics["selection_rate_disparity"]
                            if abs(disparity) > 0.2:
                                details += f"{attr}: {abs(disparity):.3f} disparity\n"

            hazard_sheet.append(
                [
                    str(run_id),
                    timestamp,
                    scores["overall"],
                    hz["id"],
                    hz["description"],
                    hz["severity"].upper(),
                    hz["mitigation"],
                    details,
                ]
            )
    else:
        # If no hazards, add a single "all clear" record
        ws.append(
            [
                str(run_id),
                timestamp,
                scores["overall"],
                scores["risk_auc"],
                scores["risk_bias"],
                "LOW",
                "COMPLETED",
                "No hazards identified",
                "N/A",
            ]
        )

    # Save locally for Streamlit dashboard
    wb.save(wb_path)

    # Save risk register to Modal Volume
    save_artifact_to_modal(
        artifact=wb,
        artifact_path=RISK_REGISTER_PATH,
    )

    print(f"✅ Risk register saved to: {RISK_REGISTER_PATH}")

    # Log metadata including hazards
    result = {
        **scores,
        "hazards": hazards,
        "risk_register_path": str(RISK_REGISTER_PATH),
    }

    log_metadata(metadata=result)
    return result
