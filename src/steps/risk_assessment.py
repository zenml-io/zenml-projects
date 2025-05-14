# Apache Software License 2.0
#
# Copyright (c) ZenML GmbH 2025. All rights reserved.
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
# http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
#

import tempfile
from pathlib import Path
from typing import Annotated, Dict, List

import pandas as pd
from openpyxl import Workbook, load_workbook
from zenml import get_step_context, log_metadata, step

from src.constants import (
    HAZARD_DEFINITIONS,
    MODAL_RISK_DIR,
    MODAL_RISK_REGISTER_PATH,
    RISK_SCORES_NAME,
)
from src.utils.modal_utils import save_artifact_to_modal

# --------------------------------------------------------------------------- #
RiskScores = Annotated[Dict[str, float], RISK_SCORES_NAME]
# --------------------------------------------------------------------------- #


def identify_hazards(evaluation_results: Dict, scores: Dict) -> List[Dict]:
    """Identify applicable hazards based on evaluation results."""
    identified_hazards = []

    # Add evaluation results to enriched context for hazard identification
    context = evaluation_results.copy()
    context.update({"scores": scores})

    # Evaluate each hazard trigger function
    for hazard_id, hazard_info in HAZARD_DEFINITIONS.items():
        try:
            if hazard_info["trigger"](context):
                identified_hazards.append(
                    {
                        "id": hazard_id,
                        "description": hazard_info["description"],
                        "severity": hazard_info["severity"],
                        "mitigation": hazard_info["mitigation"],
                    }
                )
        except (KeyError, TypeError) as e:
            # Handle missing keys gracefully
            print(f"Warning: Could not evaluate hazard {hazard_id}: {e}")

    return identified_hazards


def score_risk(evaluation: Dict) -> Dict[str, float]:
    """Return dict with per-factor risk ∈ [0,1] + overall."""
    # Example heuristics (tune for real use‑case)
    auc = evaluation["metrics"]["auc"]
    bias_flag = evaluation["fairness"].get("bias_flag", False)

    disparities = []
    # The fairness metrics are in fairness_metrics inside the fairness report
    for group_metrics in evaluation["fairness"].get("fairness_metrics", {}).values():
        if isinstance(group_metrics, dict) and "selection_rate_disparity" in group_metrics:
            disparities.append(abs(group_metrics["selection_rate_disparity"]))

    disparity = max(disparities) if disparities else 0.0

    risk_auc = 1 - auc  # low AUC → higher risk
    risk_bias = 0.8 if bias_flag else disparity  # flat score if flagged
    overall = round(min(1.0, 0.5 * risk_auc + 0.5 * risk_bias), 3)

    return {
        "risk_auc": round(risk_auc, 3),
        "risk_bias": round(risk_bias, 3),
        "overall": overall,
    }


@step
def risk_assessment(evaluation_results: Dict) -> RiskScores:
    """Compute risk scores & update register. Article 9 compliant.

    Converts evaluation metrics + bias flag → quantitative risk score
    Updates risk_register.xlsx (one row / model version)
    Logs summary via `log_metadata`

    Args:
        evaluation_results: Dictionary containing evaluation results.

    Returns:
        Dictionary containing risk scores.
    """
    scores = score_risk(evaluation_results)
    hazards = identify_hazards(evaluation_results, scores)

    # Build or update the workbook in memory
    wb_path = Path(tempfile.mkdtemp()) / "risk_register.xlsx"
    if not wb_path.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = "Risks"
        ws.append(
            ["Run_ID", "Risk_overall", "Risk_auc", "Risk_bias", "Status", "Hazards", "Mitigations"]
        )
    else:
        wb = load_workbook(wb_path)
        ws = wb["Risks"]

    # Get run_id from step context
    run_id = get_step_context().pipeline_run.id

    # Check if this run already logged
    run_ids = [cell.value for cell in ws["A"][1:]]  # skip header
    if run_id in run_ids:
        row_idx = run_ids.index(run_id) + 2
    else:
        row_idx = ws.max_row + 1

    hazard_descriptions = [h["description"] for h in hazards]
    mitigation_descriptions = [h["mitigation"] for h in hazards]

    ws.cell(row=row_idx, column=1).value = str(run_id)
    ws.cell(row=row_idx, column=2).value = scores["overall"]
    ws.cell(row=row_idx, column=3).value = scores["risk_auc"]
    ws.cell(row=row_idx, column=4).value = scores["risk_bias"]
    ws.cell(row=row_idx, column=5).value = (
        "Mitigation needed" if scores["overall"] > 0.4 else "Acceptable"
    )
    ws.cell(row=row_idx, column=6).value = "; ".join(hazard_descriptions) if hazards else "None"
    ws.cell(row=row_idx, column=7).value = "; ".join(mitigation_descriptions) if hazards else "N/A"
    wb.save(wb_path)  # write into /tmp/risk_register.xlsx

    # Save risk register to Modal Volume
    save_artifact_to_modal(
        artifact=wb,
        artifact_path=MODAL_RISK_REGISTER_PATH,
    )

    if hazards:
        # Create a hazard report with more detail
        hazard_report = f"# Risk Assessment - Run {run_id}\n\n"
        hazard_report += f"## Overall Risk Score: {scores['overall']:.2f}\n\n"
        hazard_report += "## Identified Hazards\n\n"

        for hazard in hazards:
            hazard_report += f"### {hazard['description']} (Severity: {hazard['severity']})\n"
            hazard_report += f"**Mitigation Strategy:** {hazard['mitigation']}\n\n"

        save_artifact_to_modal(
            artifact=hazard_report,
            artifact_path=f"{MODAL_RISK_DIR}/hazard_report_{run_id}.md",
        )

    # 3) (Optional) snapshot as Markdown and upload
    df = pd.read_excel(wb_path, sheet_name="Risks")
    md = df.to_markdown(index=False)
    save_artifact_to_modal(
        artifact=md,
        artifact_path=f"{MODAL_RISK_DIR}/risk_register.md",
    )

    # Log metadata including hazards
    log_metadata(
        metadata={
            "risk_scores": scores,
            "identified_hazards": hazards,
        }
    )

    # Include hazards in the return value
    scores["hazards"] = hazards
    return scores
