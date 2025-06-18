# Apache Software License 2.0
#
# Copyright (c) ZenML GmbH 2025. All rights reserved.
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
# http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
#

from datetime import datetime
from pathlib import Path
from typing import Annotated, Dict, List, Tuple

from openpyxl import Workbook, load_workbook
from zenml import get_step_context, log_metadata, step
from zenml.logger import get_logger
from zenml.types import HTMLString

from src.constants import Artifacts as A
from src.constants import Hazards
from src.utils.storage import save_artifact_to_modal

logger = get_logger(__name__)


def identify_hazards(evaluation_results: Dict, scores: Dict) -> List[Dict]:
    """Identify applicable hazards based on evaluation results."""
    identified_hazards = []

    # Get all hazards and check which ones are triggered
    hazards = Hazards.get_all()
    for hazard_id, hazard in hazards.items():
        try:
            if hazard.is_triggered(evaluation_results, scores):
                identified_hazards.append(
                    {
                        "id": hazard_id,
                        "description": hazard.description,
                        "severity": hazard.severity,
                        "mitigation": hazard.mitigation,
                    }
                )
        except (KeyError, TypeError) as e:
            logger.warning(f"Could not evaluate hazard {hazard_id}: {e}")

    return identified_hazards


def score_risk(evaluation: Dict) -> Dict[str, float]:
    """Return dict with per-factor risk ∈ [0,1] + overall."""
    # Get metrics with default fallbacks
    metrics = evaluation.get("metrics", {})
    if not isinstance(metrics, dict):
        metrics = {}

    # Get AUC with default
    auc = metrics.get("auc", 0.5)

    # Get fairness info with defaults
    fairness = evaluation.get("fairness", {})
    if not isinstance(fairness, dict):
        fairness = {}

    bias_flag = fairness.get("bias_flag", False)

    # Calculate minimum disparate impact ratio across groups
    # Lower DI ratios indicate higher bias risk (DI < 0.8 is adverse impact)
    di_ratios = []
    fairness_metrics = fairness.get("fairness_metrics", {})
    if isinstance(fairness_metrics, dict):
        for group_metrics in fairness_metrics.values():
            if (
                isinstance(group_metrics, dict)
                and "disparate_impact_ratio" in group_metrics
            ):
                di_ratios.append(group_metrics["disparate_impact_ratio"])

    min_di_ratio = min(di_ratios) if di_ratios else 1.0

    # Calculate risk scores
    risk_auc = 1 - auc  # low AUC → higher risk
    # Risk increases as DI ratio drops below 0.8 (adverse impact threshold)
    # DI ratio 1.0 = no bias risk, DI ratio 0.0 = maximum bias risk
    if bias_flag:
        risk_bias = 0.8  # High risk when bias is explicitly flagged
    else:
        # Convert DI ratio to risk score: 1.0 DI ratio = 0 risk, 0.8 DI ratio = 0.25 risk
        risk_bias = max(0.0, (0.8 - min_di_ratio) / 0.8)

    overall = round(min(1.0, 0.5 * risk_auc + 0.5 * risk_bias), 3)

    return {
        "risk_auc": round(risk_auc, 3),
        "risk_bias": round(risk_bias, 3),
        "overall": overall,
    }


def get_status(
    risk_score: float, approval_thresholds: Dict[str, float]
) -> str:
    """Determine row status for the risk register.

    Args:
        risk_score: The overall risk score ∈ [0,1].
        approval_thresholds: Dictionary containing approval thresholds.

    Returns:
        "PENDING" if the score exceeds the configured threshold, otherwise "COMPLETED".
    """
    threshold = approval_thresholds["risk_score"]
    return "PENDING" if risk_score > threshold else "COMPLETED"


def get_article_for_hazard(hazard_id: str) -> str:
    """Map hazards to specific EU AI Act articles."""
    article_mapping = {
        "BIAS_PROTECTED_GROUPS": "article_10",  # Data and Data Governance
        "poor_performance": "article_15",  # Accuracy, Robustness and Cybersecurity
        "data_quality_issues": "article_10",  # Data and Data Governance
        "model_drift": "article_17",  # Post-market Monitoring
        "security_vulnerability": "article_15",  # Accuracy, Robustness and Cybersecurity
        "transparency_issues": "article_13",  # Transparency and Provision of Information
        "human_oversight_failure": "article_14",  # Human Oversight
        "documentation_gaps": "article_11",  # Technical Documentation
    }
    return article_mapping.get(
        hazard_id, "article_9"
    )  # Default to Risk Management


def generate_risk_visualization(risk_scores: Dict, run_id: str) -> HTMLString:
    """Generate HTML visualization for risk assessment results."""
    overall_risk = risk_scores.get("overall", 0.0)
    auc_risk = risk_scores.get("risk_auc", 0.0)
    bias_risk = risk_scores.get("risk_bias", 0.0)
    hazards = risk_scores.get("hazards", [])

    # Risk level categorization
    if overall_risk < 0.3:
        risk_level = "LOW"
        risk_color = "#28a745"
        risk_bg = "#d4edda"
    elif overall_risk < 0.7:
        risk_level = "MEDIUM"
        risk_color = "#ffc107"
        risk_bg = "#fff3cd"
    else:
        risk_level = "HIGH"
        risk_color = "#dc3545"
        risk_bg = "#f8d7da"

    # Severity color mapping
    severity_colors = {
        "low": "#28a745",
        "medium": "#ffc107",
        "high": "#dc3545",
        "critical": "#6f42c1",
    }

    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <title>Risk Assessment Report - {run_id}</title>
        <style>
            body {{
                font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
                line-height: 1.6;
                color: #333;
                max-width: 1000px;
                margin: 0 auto;
                padding: 20px;
                background-color: #f8f9fa;
            }}
            .header {{
                text-align: center;
                margin-bottom: 30px;
                padding: 20px;
                background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                color: white;
                border-radius: 10px;
                box-shadow: 0 4px 6px rgba(0,0,0,0.1);
            }}
            .risk-overview {{
                display: grid;
                grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
                gap: 20px;
                margin-bottom: 30px;
            }}
            .risk-card {{
                background: white;
                padding: 20px;
                border-radius: 8px;
                box-shadow: 0 2px 4px rgba(0,0,0,0.1);
                text-align: center;
            }}
            .risk-score {{
                font-size: 2.5em;
                font-weight: bold;
                margin: 10px 0;
            }}
            .overall-risk {{
                background: {risk_bg};
                border-left: 5px solid {risk_color};
            }}
            .overall-risk .risk-score {{
                color: {risk_color};
            }}
            .risk-level {{
                font-size: 1.2em;
                font-weight: bold;
                color: {risk_color};
                margin-top: 10px;
            }}
            .hazards-section {{
                background: white;
                padding: 20px;
                border-radius: 8px;
                box-shadow: 0 2px 4px rgba(0,0,0,0.1);
                margin-bottom: 20px;
            }}
            .hazard-item {{
                border-left: 4px solid #ddd;
                padding: 15px;
                margin: 10px 0;
                background: #f8f9fa;
                border-radius: 0 5px 5px 0;
            }}
            .hazard-high {{
                border-left-color: #dc3545;
                background: #fff5f5;
            }}
            .hazard-medium {{
                border-left-color: #ffc107;
                background: #fffbf0;
            }}
            .hazard-low {{
                border-left-color: #28a745;
                background: #f0fff4;
            }}
            .hazard-id {{
                font-weight: bold;
                color: #495057;
                margin-bottom: 5px;
            }}
            .hazard-description {{
                margin-bottom: 10px;
                color: #6c757d;
            }}
            .hazard-mitigation {{
                font-style: italic;
                color: #495057;
                border-top: 1px solid #dee2e6;
                padding-top: 10px;
            }}
            .severity-badge {{
                display: inline-block;
                padding: 4px 12px;
                border-radius: 20px;
                font-size: 0.8em;
                font-weight: bold;
                text-transform: uppercase;
                color: white;
                margin-bottom: 10px;
            }}
            .no-hazards {{
                text-align: center;
                padding: 40px;
                color: #28a745;
                background: #d4edda;
                border-radius: 8px;
                border: 2px solid #28a745;
            }}
            .timestamp {{
                text-align: center;
                color: #6c757d;
                font-size: 0.9em;
                margin-top: 20px;
            }}
        </style>
    </head>
    <body>
        <div class="header">
            <h1>🛡️ Risk Assessment Report</h1>
            <p>EU AI Act Article 9 Compliance</p>
            <p><strong>Run ID:</strong> {run_id}</p>
        </div>

        <div class="risk-overview">
            <div class="risk-card overall-risk">
                <h3>Overall Risk</h3>
                <div class="risk-score">{overall_risk:.2f}</div>
                <div class="risk-level">{risk_level}</div>
            </div>
            <div class="risk-card">
                <h3>Model Performance Risk</h3>
                <div class="risk-score" style="color: #6c757d;">{auc_risk:.2f}</div>
                <small>Based on AUC Score</small>
            </div>
            <div class="risk-card">
                <h3>Bias Risk</h3>
                <div class="risk-score" style="color: #6c757d;">{bias_risk:.2f}</div>
                <small>Fairness Assessment</small>
            </div>
        </div>

        <div class="hazards-section">
            <h2>📋 Identified Hazards</h2>
            {generate_hazards_html(hazards) if hazards else '<div class="no-hazards"><h3>✅ No Hazards Identified</h3><p>The model meets all risk thresholds for this assessment.</p></div>'}
        </div>

        <div class="timestamp">
            Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S UTC')}
        </div>
    </body>
    </html>
    """

    return HTMLString(html_content)


def generate_hazards_html(hazards: List[Dict]) -> str:
    """Generate HTML for hazards list."""
    html = ""
    for hazard in hazards:
        severity = hazard.get("severity", "low").lower()
        severity_color = {
            "low": "#28a745",
            "medium": "#ffc107",
            "high": "#dc3545",
            "critical": "#6f42c1",
        }.get(severity, "#6c757d")

        hazard_class = f"hazard-{severity}"

        html += f"""
        <div class="hazard-item {hazard_class}">
            <div class="hazard-id">{hazard.get('id', 'UNKNOWN')}</div>
            <div class="severity-badge" style="background-color: {severity_color};">
                {severity.upper()}
            </div>
            <div class="hazard-description">{hazard.get('description', 'No description available')}</div>
            <div class="hazard-mitigation">
                <strong>Mitigation:</strong> {hazard.get('mitigation', 'No mitigation specified')}
            </div>
        </div>
        """

    return html


@step
def risk_assessment(
    evaluation_results: Dict,
    approval_thresholds: Dict[str, float],
    risk_register_path: str = "docs/risk/risk_register.xlsx",
) -> Tuple[
    Annotated[Dict, A.RISK_SCORES], Annotated[HTMLString, A.RISK_VISUALIZATION]
]:
    """Compute risk scores & update register. Article 9 compliant."""
    scores = score_risk(evaluation_results)
    hazards = identify_hazards(evaluation_results, scores)

    headers = [
        "Run_ID",
        "Timestamp",
        "Risk_overall",
        "Risk_auc",
        "Risk_bias",
        "Risk_category",
        "Status",
        "Risk_description",
        "Mitigation",
        "Article",
        "Mitigation_status",
        "Review_date",
    ]

    # Load or create workbook
    wb_path = Path(risk_register_path)
    if wb_path.exists():
        wb = load_workbook(wb_path)
        if "Risks" in wb.sheetnames:
            ws = wb["Risks"]
        else:
            ws = wb.create_sheet("Risks")
            ws.append(headers)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Risks"
        ws.append(headers)

    run_id = get_step_context().pipeline_run.id
    timestamp = datetime.now().isoformat()

    # Process hazards
    if hazards:
        for hz in hazards:
            status = get_status(scores["overall"], approval_thresholds)
            article = get_article_for_hazard(hz["id"])

            ws.append(
                [
                    str(run_id),
                    timestamp,
                    scores["overall"],
                    scores["risk_auc"],
                    scores["risk_bias"],
                    hz["severity"].upper(),
                    status,
                    hz["description"],
                    hz["mitigation"],
                    article,
                    status,
                    timestamp,
                ]
            )

        # HazardDetails sheet
        if "HazardDetails" not in wb.sheetnames:
            hazard_sheet = wb.create_sheet("HazardDetails")
            hazard_sheet.append(
                [
                    "Run_ID",
                    "Timestamp",
                    "Risk_Score",
                    "Hazard_ID",
                    "Description",
                    "Severity",
                    "Mitigation",
                    "Details",
                    "Article",
                ]
            )
        else:
            hazard_sheet = wb["HazardDetails"]

        for hz in hazards:
            details = ""
            if (
                hz["id"] == "BIAS_PROTECTED_GROUPS"
                and "fairness" in evaluation_results
            ):
                fairness = evaluation_results["fairness"]
                if (
                    isinstance(fairness, dict)
                    and "fairness_metrics" in fairness
                ):
                    for attr, metrics in fairness.get(
                        "fairness_metrics", {}
                    ).items():
                        if (
                            isinstance(metrics, dict)
                            and "selection_rate_disparity" in metrics
                        ):
                            di_ratio = metrics.get(
                                "disparate_impact_ratio", 1.0
                            )
                            if di_ratio < 0.8:  # Adverse impact threshold
                                details += f"{attr}: {di_ratio:.3f} DI ratio (< 0.8 indicates adverse impact)\n"

            article = get_article_for_hazard(hz["id"])
            hazard_sheet.append(
                [
                    str(run_id),
                    timestamp,
                    scores["overall"],
                    hz["id"],
                    hz["description"],
                    hz["severity"].upper(),
                    hz["mitigation"],
                    details,
                    article,
                ]
            )
    else:
        # No hazards case
        status = get_status(scores["overall"], approval_thresholds)
        ws.append(
            [
                str(run_id),
                timestamp,
                scores["overall"],
                scores["risk_auc"],
                scores["risk_bias"],
                "LOW",
                status,
                "No hazards identified",
                "N/A",
                "article_9",
                status,
                timestamp,
            ]
        )

    wb.save(wb_path)
    save_artifact_to_modal(artifact=wb, artifact_path=risk_register_path)

    result = {
        **scores,
        "hazards": hazards,
        "risk_register_path": str(risk_register_path),
    }
    log_metadata(metadata=result)

    # Generate visualization
    run_id = get_step_context().pipeline_run.id
    risk_visualization = generate_risk_visualization(result, str(run_id))

    return result, risk_visualization
